"""
Excel Rapport Generator - Wekelijks Vastgoedrapport
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint
from datetime import datetime
import os

# ── Kleurenpalet ──────────────────────────────────────────────────────────────
C_DARK_BLUE  = "1B3A5C"
C_MID_BLUE   = "2E6DA4"
C_LIGHT_BLUE = "D6E4F0"
C_ACCENT     = "E8A020"
C_GREEN      = "217346"
C_LIGHT_GRN  = "E2EFDA"
C_ORANGE     = "C55A11"
C_LIGHT_ORG  = "FCE4D6"
C_PURPLE     = "5B2D8E"
C_LIGHT_PUR  = "EAE0F5"
C_GRAY_DARK  = "404040"
C_GRAY_MED   = "A6A6A6"
C_GRAY_LIGHT = "F2F2F2"
C_WHITE      = "FFFFFF"

EURO_FMT  = '#,##0\ "€"'
M2_FMT    = '#,##0\ "m²"'
EURO_M2   = '#,##0\ "€/m²"'
PCT_FMT   = "0.0%"


def thick_border():
    s = Side(style="medium", color=C_DARK_BLUE)
    return Border(left=s, right=s, top=s, bottom=s)


def thin_border():
    s = Side(style="thin", color=C_GRAY_MED)
    return Border(left=s, right=s, top=s, bottom=s)


def header_cell(ws, row, col, value, bg=C_DARK_BLUE, fg=C_WHITE, size=11, bold=True, wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Arial", bold=bold, color=fg, size=size)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center",
                             wrap_text=wrap)
    c.border = thin_border()
    return c


def data_cell(ws, row, col, value, fmt=None, bg=None, bold=False, align="left"):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Arial", size=10, bold=bold,
                  color=C_GRAY_DARK if not bold else C_DARK_BLUE)
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    if fmt:
        c.number_format = fmt
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border = thin_border()
    return c


def set_col_widths(ws, widths):
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w


def add_title_block(ws, title, subtitle, row_start=1):
    ws.row_dimensions[row_start].height = 40
    ws.row_dimensions[row_start + 1].height = 22

    # Merge title across all columns
    ws.merge_cells(start_row=row_start, start_column=1,
                   end_row=row_start, end_column=12)
    t = ws.cell(row=row_start, column=1, value=title)
    t.font = Font(name="Arial", bold=True, size=18, color=C_WHITE)
    t.fill = PatternFill("solid", fgColor=C_DARK_BLUE)
    t.alignment = Alignment(horizontal="left", vertical="center",
                             indent=1)

    ws.merge_cells(start_row=row_start + 1, start_column=1,
                   end_row=row_start + 1, end_column=12)
    s = ws.cell(row=row_start + 1, column=1, value=subtitle)
    s.font = Font(name="Arial", italic=True, size=11, color=C_WHITE)
    s.fill = PatternFill("solid", fgColor=C_MID_BLUE)
    s.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    return row_start + 2


def build_listing_sheet(wb, sheet_name, data, tab_color):
    """Bouwt een gedetailleerde listing-sheet."""
    ws = wb.create_sheet(sheet_name)
    ws.sheet_properties.tabColor = tab_color
    ws.freeze_panes = "A5"

    week = datetime.now().strftime("Week %W — %d %B %Y")
    row = add_title_block(ws, f"📋 {sheet_name}", week)

    # Kolomhoofdingen
    headers = [
        "Provincie", "Gemeente", "Adres", "Postcode",
        "Opp. (m²)", "Bouwjaar", "BTW / Regime",
        "Vraagprijs", "Prijs/m²", "Type", "Bron", "Datum"
    ]
    ws.row_dimensions[row].height = 36
    for c, h in enumerate(headers, 1):
        header_cell(ws, row, c, h, bg=C_MID_BLUE, wrap=True)
    row += 1

    # Data
    for i, rec in enumerate(data):
        bg = C_WHITE if i % 2 == 0 else C_GRAY_LIGHT
        data_cell(ws, row, 1,  rec.get("provincie", ""), bg=bg)
        data_cell(ws, row, 2,  rec.get("gemeente", ""), bg=bg)
        data_cell(ws, row, 3,  rec.get("adres", ""), bg=bg)
        data_cell(ws, row, 4,  rec.get("postcode", ""), bg=bg, align="center")
        data_cell(ws, row, 5,  rec.get("oppervlakte_m2"), fmt=M2_FMT, bg=bg, align="right")
        data_cell(ws, row, 6,  rec.get("bouwjaar"), bg=bg, align="center")
        # BTW-regime kleur
        btw = rec.get("btw_regime", "")
        btw_bg = bg
        if "21%" in btw:
            btw_bg = C_LIGHT_ORG
        elif "6%" in btw:
            btw_bg = C_LIGHT_GRN
        data_cell(ws, row, 7,  btw, bg=btw_bg)
        data_cell(ws, row, 8,  rec.get("prijs"), fmt=EURO_FMT, bg=bg, align="right")
        data_cell(ws, row, 9,  rec.get("prijs_per_m2"), fmt=EURO_M2, bg=bg, align="right")
        data_cell(ws, row, 10, rec.get("type", ""), bg=bg)
        data_cell(ws, row, 11, rec.get("bron", ""), bg=bg)
        data_cell(ws, row, 12, rec.get("scraped_op", ""), bg=bg, align="center")
        row += 1

    set_col_widths(ws, [16, 16, 32, 10, 11, 10, 22, 16, 13, 14, 18, 13])

    # AutoFilter
    ws.auto_filter.ref = f"A4:{get_column_letter(len(headers))}{row - 1}"
    return ws


def build_summary_sheet(wb, all_data):
    """Bouwt het samenvatting-dashboard."""
    ws = wb.create_sheet("📊 Samenvatting", 0)
    ws.sheet_properties.tabColor = C_DARK_BLUE

    week = datetime.now().strftime("Week %W — %d %B %Y")
    row = add_title_block(
        ws,
        "🏠 Wekelijks Vastgoedmarkt Rapport",
        f"Provincies: Antwerpen · Vlaams-Brabant · Limburg  |  {week}  |  Bron: Immoweb & Zimmo"
    )
    row += 1  # witregel

    provincies = ["Antwerpen", "Vlaams-Brabant", "Limburg"]
    btw_regimes = ["Nieuwbouw 21% BTW", "Nieuwbouw 6% BTW", "Bestaand (reg. rechten)"]
    types_main = ["Woning", "Appartement"]
    alle_types = ["Woning", "Appartement", "Klein (<50m²)"]

    # ── Sectie 1: Statistieken per provincie ──────────────────────────────────
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    sec = ws.cell(row=row, column=1, value="📍 Gemiddelde Prijs/m² per Provincie & Type")
    sec.font = Font(name="Arial", bold=True, size=13, color=C_WHITE)
    sec.fill = PatternFill("solid", fgColor=C_DARK_BLUE)
    sec.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 28
    row += 1

    # Header rij
    col_headers = ["Provincie"] + alle_types + ["Totaal", "Mediaan", "# Listings"]
    for c, h in enumerate(col_headers, 1):
        header_cell(ws, row, c, h, bg=C_MID_BLUE)
    ws.row_dimensions[row].height = 30
    row += 1

    prov_stats = {}
    for prov in provincies:
        prov_data = [r for r in all_data if r["provincie"] == prov]
        row_data = {"Provincie": prov}
        all_pm2 = []
        for t in alle_types:
            subset = [r for r in prov_data if r["type"] == t and r.get("prijs_per_m2")]
            pm2 = [r["prijs_per_m2"] for r in subset]
            row_data[t] = round(sum(pm2) / len(pm2)) if pm2 else None
            all_pm2.extend(pm2)
        row_data["Totaal"] = round(sum(all_pm2) / len(all_pm2)) if all_pm2 else None
        row_data["Mediaan"] = sorted(all_pm2)[len(all_pm2) // 2] if all_pm2 else None
        row_data["# Listings"] = len(prov_data)
        prov_stats[prov] = row_data

        bg = C_LIGHT_BLUE
        data_cell(ws, row, 1, prov, bg=bg, bold=True)
        for ci, t in enumerate(alle_types, 2):
            v = row_data.get(t)
            data_cell(ws, row, ci, v, fmt=EURO_M2, bg=bg, align="right")
        data_cell(ws, row, 5, row_data["Totaal"], fmt=EURO_M2, bg=bg, bold=True, align="right")
        data_cell(ws, row, 6, row_data["Mediaan"], fmt=EURO_M2, bg=bg, align="right")
        data_cell(ws, row, 7, row_data["# Listings"], bg=bg, align="center")
        row += 1

    set_col_widths(ws, [20, 16, 16, 16, 16, 16, 14, 14, 14, 14])
    row += 1

    # ── Sectie 2: BTW-analyse ─────────────────────────────────────────────────
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    sec2 = ws.cell(row=row, column=1, value="🏗️ Analyse per BTW-Regime")
    sec2.font = Font(name="Arial", bold=True, size=13, color=C_WHITE)
    sec2.fill = PatternFill("solid", fgColor=C_DARK_BLUE)
    sec2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 28
    row += 1

    btw_col_headers = ["BTW Regime"] + provincies + ["Gem. Prijs/m²", "# Listings"]
    for c, h in enumerate(btw_col_headers, 1):
        header_cell(ws, row, c, h, bg=C_MID_BLUE)
    ws.row_dimensions[row].height = 30
    row += 1

    btw_colors = {
        "Nieuwbouw 21% BTW":       C_LIGHT_ORG,
        "Nieuwbouw 6% BTW":        C_LIGHT_GRN,
        "Bestaand (reg. rechten)": C_LIGHT_PUR,
    }

    for btw in btw_regimes:
        btw_subset = [r for r in all_data if r["btw_regime"] == btw]
        bg = btw_colors.get(btw, C_GRAY_LIGHT)
        data_cell(ws, row, 1, btw, bg=bg, bold=True)
        for ci, prov in enumerate(provincies, 2):
            sub = [r for r in btw_subset if r["provincie"] == prov and r.get("prijs_per_m2")]
            pm2 = [r["prijs_per_m2"] for r in sub]
            v = round(sum(pm2) / len(pm2)) if pm2 else None
            data_cell(ws, row, ci, v, fmt=EURO_M2, bg=bg, align="right")
        all_pm2 = [r["prijs_per_m2"] for r in btw_subset if r.get("prijs_per_m2")]
        data_cell(ws, row, 5, round(sum(all_pm2) / len(all_pm2)) if all_pm2 else None,
                  fmt=EURO_M2, bg=bg, bold=True, align="right")
        data_cell(ws, row, 6, len(btw_subset), bg=bg, align="center")
        row += 1

    row += 1

    # ── Sectie 3: Top 10 duurste postcodergebieden ────────────────────────────
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    sec3 = ws.cell(row=row, column=1, value="📈 Top 10 Duurste Postcodes (prijs/m²)")
    sec3.font = Font(name="Arial", bold=True, size=13, color=C_WHITE)
    sec3.fill = PatternFill("solid", fgColor=C_DARK_BLUE)
    sec3.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 28
    row += 1

    top_col_h = ["Rang", "Postcode", "Gemeente", "Provincie", "Gem. Prijs/m²", "# Listings"]
    for c, h in enumerate(top_col_h, 1):
        header_cell(ws, row, c, h, bg=C_MID_BLUE)
    ws.row_dimensions[row].height = 30
    row += 1

    # Bereken gemiddelde per postcode
    pc_stats = {}
    for r_data in all_data:
        pc = r_data.get("postcode", "0000")
        pm2 = r_data.get("prijs_per_m2")
        if pm2:
            if pc not in pc_stats:
                pc_stats[pc] = {
                    "gemeente": r_data.get("gemeente", ""),
                    "provincie": r_data.get("provincie", ""),
                    "pm2_list": []
                }
            pc_stats[pc]["pm2_list"].append(pm2)

    top10 = sorted(
        [(pc, d) for pc, d in pc_stats.items() if len(d["pm2_list"]) >= 3],
        key=lambda x: sum(x[1]["pm2_list"]) / len(x[1]["pm2_list"]),
        reverse=True
    )[:10]

    medal_colors = [C_ACCENT, C_GRAY_MED, "C87941"]  # goud, zilver, brons
    for rank, (pc, d) in enumerate(top10, 1):
        gem = round(sum(d["pm2_list"]) / len(d["pm2_list"]))
        bg = medal_colors[rank - 1] if rank <= 3 else (C_WHITE if rank % 2 else C_GRAY_LIGHT)
        fg_bold = rank <= 3
        data_cell(ws, row, 1, rank, bg=bg, bold=fg_bold, align="center")
        data_cell(ws, row, 2, pc, bg=bg, bold=fg_bold, align="center")
        data_cell(ws, row, 3, d["gemeente"], bg=bg, bold=fg_bold)
        data_cell(ws, row, 4, d["provincie"], bg=bg)
        data_cell(ws, row, 5, gem, fmt=EURO_M2, bg=bg, bold=fg_bold, align="right")
        data_cell(ws, row, 6, len(d["pm2_list"]), bg=bg, align="center")
        row += 1

    row += 2

    # ── Legenda ───────────────────────────────────────────────────────────────
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    leg = ws.cell(row=row, column=1, value="💡 Legenda BTW-Regime")
    leg.font = Font(name="Arial", bold=True, size=11, color=C_DARK_BLUE)
    leg.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 22
    row += 1

    legend_items = [
        (C_LIGHT_ORG, "Nieuwbouw 21% BTW",       "Nieuwe constructies, standaard BTW-tarief"),
        (C_LIGHT_GRN, "Nieuwbouw 6% BTW",         "Renovatie of sociale woningbouw, verlaagd BTW-tarief"),
        (C_LIGHT_PUR, "Bestaand (reg. rechten)",  "Bestaand vastgoed, registratierechten van toepassing"),
    ]
    for bg, label, uitleg in legend_items:
        c1 = ws.cell(row=row, column=1, value=label)
        c1.font = Font(name="Arial", bold=True, size=10)
        c1.fill = PatternFill("solid", fgColor=bg)
        c1.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c1.border = thin_border()
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        c2 = ws.cell(row=row, column=2, value=uitleg)
        c2.font = Font(name="Arial", size=10)
        c2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c2.border = thin_border()
        ws.row_dimensions[row].height = 20
        row += 1

    return ws


def generate_report(data, output_path="vastgoed_rapport.xlsx"):
    wb = openpyxl.Workbook()
    # Verwijder standaard sheet
    wb.remove(wb.active)

    # ── Filters ───────────────────────────────────────────────────────────────
    woningen     = [r for r in data if r["type"] == "Woning"]
    appartementen = [r for r in data if r["type"] == "Appartement"]
    kleine        = [r for r in data if r["type"] == "Klein (<50m²)"]

    nieuwbouw_21 = [r for r in data if r["btw_regime"] == "Nieuwbouw 21% BTW"]
    nieuwbouw_6  = [r for r in data if r["btw_regime"] == "Nieuwbouw 6% BTW"]
    bestaand     = [r for r in data if r["btw_regime"] == "Bestaand (reg. rechten)"]

    # ── Sheets aanmaken ───────────────────────────────────────────────────────
    build_summary_sheet(wb, data)

    build_listing_sheet(wb, "🏡 Woningen — Bestaand",
                        [r for r in woningen if r["btw_regime"] == "Bestaand (reg. rechten)"],
                        "217346")
    build_listing_sheet(wb, "🏡 Woningen — Nieuwbouw",
                        [r for r in woningen if "Nieuwbouw" in r["btw_regime"]],
                        "C55A11")
    build_listing_sheet(wb, "🏢 Appartementen — Bestaand",
                        [r for r in appartementen if r["btw_regime"] == "Bestaand (reg. rechten)"],
                        "2E6DA4")
    build_listing_sheet(wb, "🏢 Appartementen — Nieuwbouw",
                        [r for r in appartementen if "Nieuwbouw" in r["btw_regime"]],
                        "5B2D8E")
    build_listing_sheet(wb, "📐 Klein (<50m²)",
                        kleine, "BF9000")
    build_listing_sheet(wb, "🏗️ Nieuwbouw 21% BTW",
                        nieuwbouw_21, "FF0000")
    build_listing_sheet(wb, "♻️ Nieuwbouw 6% BTW",
                        nieuwbouw_6,  "70AD47")
    build_listing_sheet(wb, "📋 Alle Listings",
                        data, "404040")

    # Properties
    wb.properties.title = "Wekelijks Vastgoedrapport"
    wb.properties.creator = "Vastgoed Scraper Pro"
    wb.properties.description = (
        "Wekelijks marktrapport — Antwerpen, Vlaams-Brabant, Limburg"
    )

    wb.save(output_path)
    print(f"✅ Excel rapport opgeslagen: {output_path}")
    return output_path
